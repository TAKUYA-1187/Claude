"""
レポート生成モジュール
======================
各種フォーマットでレポートを出力:
  - Excel (複数シート: 売上/在庫/商品別/月次)
  - CSV (データエクスポート)
  - HTML (ブラウザで閲覧可能な視覚的レポート)
"""

import os
import json
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, Optional
from config import APP_CONFIG


REPORT_DIR = APP_CONFIG["report_output_dir"]
os.makedirs(REPORT_DIR, exist_ok=True)


class ReportGenerator:

    def __init__(self):
        self.currency_symbol = APP_CONFIG["currency_symbol"]
        self.generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    def _fmt_currency(self, value: float) -> str:
        return f"{self.currency_symbol}{value:,.0f}"

    def _fmt_pct(self, value: float) -> str:
        sign = "+" if value > 0 else ""
        return f"{sign}{value:.1f}%"

    # ================================================================
    # Excel レポート
    # ================================================================
    def export_excel(self, kpi: Dict, daily_df: pd.DataFrame,
                     monthly_df: pd.DataFrame, product_df: pd.DataFrame,
                     inventory_df: pd.DataFrame,
                     filename: str = None) -> str:
        """包括的Excelレポートの生成"""
        try:
            import openpyxl
            from openpyxl.styles import (PatternFill, Font, Alignment,
                                          Border, Side, numbers)
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, LineChart, Reference
        except ImportError:
            print("openpyxlが未インストールです: pip install openpyxl")
            return self.export_csv_bundle(daily_df, monthly_df, product_df, inventory_df)

        if filename is None:
            filename = f"amazon_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(REPORT_DIR, filename)

        wb = openpyxl.Workbook()

        # --- スタイル定義 ---
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        subheader_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        alt_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        good_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        warn_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        danger_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="FF8C00")
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        def style_header_row(ws, row, col_count):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # ================================================================
        # Sheet 1: KPIサマリー
        # ================================================================
        ws1 = wb.active
        ws1.title = "KPIサマリー"
        ws1.sheet_view.showGridLines = False

        ws1["A1"] = "Amazon セラーアナリティクス - KPIサマリー"
        ws1["A1"].font = title_font
        ws1["A2"] = f"生成日時: {self.generated_at}  |  セラーID: {APP_CONFIG.get('seller_id', 'A2B1UA7OXUW3IW')}"
        ws1["A2"].font = Font(italic=True, color="888888")
        ws1.row_dimensions[3].height = 10

        kpi_items = [
            ("総売上", self._fmt_currency(kpi.get("total_revenue", 0))),
            ("総注文数", f"{kpi.get('total_orders', 0):,} 件"),
            ("出荷済み注文", f"{kpi.get('shipped_orders', 0):,} 件"),
            ("保留中注文", f"{kpi.get('pending_orders', 0):,} 件"),
            ("平均注文金額", self._fmt_currency(kpi.get("avg_order_value", 0))),
            ("最大注文金額", self._fmt_currency(kpi.get("max_order", 0))),
            ("今月売上", self._fmt_currency(kpi.get("this_month_revenue", 0))),
            ("先月売上", self._fmt_currency(kpi.get("last_month_revenue", 0))),
            ("前月比成長率", self._fmt_pct(kpi.get("mom_growth_pct", 0))),
        ]

        ws1["A4"] = "KPI項目"
        ws1["B4"] = "値"
        style_header_row(ws1, 4, 2)

        for i, (label, value) in enumerate(kpi_items, start=5):
            ws1.cell(row=i, column=1, value=label).border = thin_border
            cell = ws1.cell(row=i, column=2, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            if i % 2 == 0:
                ws1.cell(row=i, column=1).fill = alt_fill
                cell.fill = alt_fill

        mom = kpi.get("mom_growth_pct", 0)
        growth_cell = ws1.cell(row=13, column=2)
        growth_cell.fill = good_fill if mom >= 0 else danger_fill
        growth_cell.font = Font(bold=True, color="2E7D32" if mom >= 0 else "C62828")

        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 20

        # ================================================================
        # Sheet 2: 月次売上
        # ================================================================
        ws2 = wb.create_sheet("月次売上")
        ws2.sheet_view.showGridLines = False
        ws2["A1"] = "月次売上推移"
        ws2["A1"].font = title_font

        if not monthly_df.empty:
            headers = ["年月", "売上", "注文数", "平均注文額", "前月比成長率"]
            for c, h in enumerate(headers, 1):
                ws2.cell(row=3, column=c, value=h)
            style_header_row(ws2, 3, len(headers))

            for i, row in enumerate(monthly_df.itertuples(), start=4):
                ws2.cell(row=i, column=1, value=str(getattr(row, "year_month", "")))
                ws2.cell(row=i, column=2, value=round(getattr(row, "revenue", 0), 0))
                ws2.cell(row=i, column=3, value=getattr(row, "orders", 0))
                ws2.cell(row=i, column=4, value=round(getattr(row, "avg_order_value", 0), 0))
                growth = getattr(row, "revenue_growth", None)
                ws2.cell(row=i, column=5, value=round(growth, 1) if pd.notna(growth) else "")

                for c in range(1, 6):
                    ws2.cell(row=i, column=c).border = thin_border
                if i % 2 == 0:
                    for c in range(1, 6):
                        ws2.cell(row=i, column=c).fill = alt_fill

        auto_width(ws2)

        # ================================================================
        # Sheet 3: 商品別ランキング
        # ================================================================
        ws3 = wb.create_sheet("商品別ランキング")
        ws3.sheet_view.showGridLines = False
        ws3["A1"] = "商品別売上ランキング"
        ws3["A1"].font = title_font

        if not product_df.empty:
            headers = ["順位", "商品名", "ASIN", "売上", "販売数", "注文数", "平均価格", "売上シェア"]
            for c, h in enumerate(headers, 1):
                ws3.cell(row=3, column=c, value=h)
            style_header_row(ws3, 3, len(headers))

            for i, row in enumerate(product_df.itertuples(), start=4):
                rank = i - 3
                ws3.cell(row=i, column=1, value=rank)
                ws3.cell(row=i, column=2, value=getattr(row, "product_name", ""))
                ws3.cell(row=i, column=3, value=getattr(row, "asin", ""))
                ws3.cell(row=i, column=4, value=round(getattr(row, "revenue", 0), 0))
                ws3.cell(row=i, column=5, value=getattr(row, "quantity", 0))
                ws3.cell(row=i, column=6, value=getattr(row, "orders", 0))
                ws3.cell(row=i, column=7, value=round(getattr(row, "avg_price", 0), 0))
                share = getattr(row, "revenue_share", 0)
                ws3.cell(row=i, column=8, value=f"{share:.1f}%")

                fill = [good_fill, good_fill, good_fill][0] if rank <= 3 else (
                    alt_fill if i % 2 == 0 else None
                )
                for c in range(1, 9):
                    ws3.cell(row=i, column=c).border = thin_border
                    if fill:
                        ws3.cell(row=i, column=c).fill = fill

        auto_width(ws3)

        # ================================================================
        # Sheet 4: 在庫状況
        # ================================================================
        ws4 = wb.create_sheet("在庫状況")
        ws4.sheet_view.showGridLines = False
        ws4["A1"] = "在庫状況一覧"
        ws4["A1"].font = title_font

        if not inventory_df.empty:
            cols = ["product_name", "seller_sku", "asin", "fulfillable_qty",
                    "inbound_working", "inbound_shipped", "reserved_total"]
            extra_cols = []
            if "stock_status" in inventory_df.columns:
                cols.append("stock_status")
            if "days_of_stock" in inventory_df.columns:
                cols.extend(["days_of_stock", "reorder_qty"])

            col_labels = {
                "product_name": "商品名", "seller_sku": "セラーSKU", "asin": "ASIN",
                "fulfillable_qty": "販売可能在庫", "inbound_working": "入荷予定",
                "inbound_shipped": "輸送中", "reserved_total": "予約済",
                "stock_status": "在庫状態", "days_of_stock": "在庫日数",
                "reorder_qty": "補充推奨数",
            }
            display_cols = [c for c in cols if c in inventory_df.columns]
            headers = [col_labels.get(c, c) for c in display_cols]

            for c, h in enumerate(headers, 1):
                ws4.cell(row=3, column=c, value=h)
            style_header_row(ws4, 3, len(headers))

            for i, row in enumerate(inventory_df[display_cols].itertuples(), start=4):
                for c, col in enumerate(display_cols, 1):
                    val = getattr(row, col, "")
                    ws4.cell(row=i, column=c, value=val)
                    ws4.cell(row=i, column=c).border = thin_border

                if "stock_status" in display_cols:
                    status_col = display_cols.index("stock_status") + 1
                    status_val = getattr(row, "stock_status", "")
                    status_cell = ws4.cell(row=i, column=status_col)
                    if status_val == "在庫切れ":
                        status_cell.fill = danger_fill
                        status_cell.font = Font(bold=True, color="C62828")
                    elif status_val in ("危険水準", "低在庫"):
                        status_cell.fill = warn_fill
                        status_cell.font = Font(bold=True, color="F57F17")
                    else:
                        status_cell.fill = good_fill

        auto_width(ws4)

        # ================================================================
        # Sheet 5: 日次売上 (生データ)
        # ================================================================
        ws5 = wb.create_sheet("日次売上")
        ws5.sheet_view.showGridLines = False
        ws5["A1"] = "日次売上データ"
        ws5["A1"].font = title_font

        if not daily_df.empty:
            headers = ["日付", "売上", "注文数", "平均注文額", "7日移動平均"]
            for c, h in enumerate(headers, 1):
                ws5.cell(row=3, column=c, value=h)
            style_header_row(ws5, 3, len(headers))

            for i, row in enumerate(daily_df.itertuples(), start=4):
                ws5.cell(row=i, column=1, value=str(getattr(row, "date", ""))[:10])
                ws5.cell(row=i, column=2, value=round(getattr(row, "revenue", 0), 0))
                ws5.cell(row=i, column=3, value=getattr(row, "orders", 0))
                ws5.cell(row=i, column=4, value=round(getattr(row, "avg_order_value", 0), 0))
                ma = getattr(row, "revenue_ma7", 0)
                ws5.cell(row=i, column=5, value=round(ma, 0) if pd.notna(ma) else "")
                for c in range(1, 6):
                    ws5.cell(row=i, column=c).border = thin_border

        auto_width(ws5)

        wb.save(filepath)
        return filepath

    # ================================================================
    # CSV バンドル
    # ================================================================
    def export_csv_bundle(self, daily_df: pd.DataFrame, monthly_df: pd.DataFrame,
                          product_df: pd.DataFrame, inventory_df: pd.DataFrame) -> str:
        """複数CSVファイルとしてエクスポート"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        files = []

        exports = [
            (daily_df, f"daily_sales_{timestamp}.csv"),
            (monthly_df, f"monthly_sales_{timestamp}.csv"),
            (product_df, f"product_ranking_{timestamp}.csv"),
            (inventory_df, f"inventory_{timestamp}.csv"),
        ]
        for df, fname in exports:
            if not df.empty:
                path = os.path.join(REPORT_DIR, fname)
                df.to_csv(path, index=False, encoding="utf-8-sig")
                files.append(path)

        return ", ".join(files)

    # ================================================================
    # HTML レポート
    # ================================================================
    def export_html(self, kpi: Dict, daily_df: pd.DataFrame,
                    monthly_df: pd.DataFrame, product_df: pd.DataFrame,
                    inventory_df: pd.DataFrame,
                    filename: str = None) -> str:
        """視覚的HTMLレポートの生成"""
        if filename is None:
            filename = f"amazon_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        filepath = os.path.join(REPORT_DIR, filename)

        mom = kpi.get("mom_growth_pct", 0)
        mom_color = "#2E7D32" if mom >= 0 else "#C62828"
        mom_arrow = "▲" if mom >= 0 else "▼"

        kpi_cards = f"""
        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-label">総売上</div>
                <div class="kpi-value">{self._fmt_currency(kpi.get('total_revenue', 0))}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">総注文数</div>
                <div class="kpi-value">{kpi.get('total_orders', 0):,} 件</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">平均注文金額</div>
                <div class="kpi-value">{self._fmt_currency(kpi.get('avg_order_value', 0))}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">今月売上</div>
                <div class="kpi-value">{self._fmt_currency(kpi.get('this_month_revenue', 0))}</div>
                <div class="kpi-sub" style="color:{mom_color}">{mom_arrow} {abs(mom):.1f}% 前月比</div>
            </div>
        </div>
        """

        # 月次テーブル
        monthly_rows = ""
        if not monthly_df.empty:
            for _, row in monthly_df.iterrows():
                growth = row.get("revenue_growth", None)
                growth_str = (f'<span style="color:{"#2E7D32" if growth >= 0 else "#C62828"}">'
                              f'{"+" if growth >= 0 else ""}{growth:.1f}%</span>'
                              ) if pd.notna(growth) else "-"
                monthly_rows += f"""
                <tr>
                    <td>{str(row.get('year_month', ''))}</td>
                    <td class="num">{self._fmt_currency(row.get('revenue', 0))}</td>
                    <td class="num">{int(row.get('orders', 0)):,}</td>
                    <td class="num">{self._fmt_currency(row.get('avg_order_value', 0))}</td>
                    <td class="num">{growth_str}</td>
                </tr>"""

        # 商品ランキングテーブル
        product_rows = ""
        if not product_df.empty:
            for rank, row in enumerate(product_df.itertuples(), 1):
                badge = (f'<span class="badge badge-gold">{rank}</span>' if rank == 1 else
                         f'<span class="badge badge-silver">{rank}</span>' if rank == 2 else
                         f'<span class="badge badge-bronze">{rank}</span>' if rank == 3 else
                         str(rank))
                product_rows += f"""
                <tr>
                    <td class="center">{badge}</td>
                    <td>{getattr(row, 'product_name', '')}</td>
                    <td class="mono">{getattr(row, 'asin', '')}</td>
                    <td class="num">{self._fmt_currency(getattr(row, 'revenue', 0))}</td>
                    <td class="num">{int(getattr(row, 'quantity', 0)):,}</td>
                    <td class="num">{getattr(row, 'revenue_share', 0):.1f}%</td>
                </tr>"""

        # 在庫テーブル
        inventory_rows = ""
        if not inventory_df.empty:
            for _, row in inventory_df.iterrows():
                status = row.get("stock_status", "不明")
                status_class = ("danger" if status == "在庫切れ" else
                                "warn" if status in ("危険水準", "低在庫") else "good")
                inventory_rows += f"""
                <tr>
                    <td>{row.get('product_name', '')}</td>
                    <td class="mono">{row.get('seller_sku', '')}</td>
                    <td class="num">{int(row.get('fulfillable_qty', 0)):,}</td>
                    <td class="num">{int(row.get('inbound_working', 0) + row.get('inbound_shipped', 0)):,}</td>
                    <td class="center"><span class="status {status_class}">{status}</span></td>
                    <td class="num">{row.get('days_of_stock', '-') if 'days_of_stock' in row.index else '-'}</td>
                    <td class="num">{int(row.get('reorder_qty', 0)) if 'reorder_qty' in row.index else '-'}</td>
                </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Amazon セラーアナリティクス レポート</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Noto Sans JP', 'Hiragino Sans', sans-serif;
         background: #f5f5f5; color: #333; }}
  .header {{ background: linear-gradient(135deg, #FF8C00, #FF6600);
             color: white; padding: 24px 32px; }}
  .header h1 {{ font-size: 24px; font-weight: 700; }}
  .header p {{ opacity: 0.85; margin-top: 4px; font-size: 13px; }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 24px 16px; }}
  .section {{ background: white; border-radius: 10px; padding: 24px;
              margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  .section h2 {{ font-size: 18px; color: #FF8C00; margin-bottom: 16px;
                 border-bottom: 2px solid #FF8C00; padding-bottom: 8px; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
               gap: 16px; }}
  .kpi-card {{ background: white; border-radius: 10px; padding: 20px;
               box-shadow: 0 2px 8px rgba(0,0,0,.08);
               border-left: 4px solid #FF8C00; }}
  .kpi-label {{ font-size: 13px; color: #888; margin-bottom: 6px; }}
  .kpi-value {{ font-size: 26px; font-weight: 700; color: #333; }}
  .kpi-sub {{ font-size: 13px; font-weight: 600; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th {{ background: #FF8C00; color: white; padding: 10px 12px; text-align: left; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #eee; }}
  tr:hover td {{ background: #FFF3E0; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .center {{ text-align: center; }}
  .mono {{ font-family: monospace; font-size: 12px; }}
  .badge {{ display: inline-block; width: 28px; height: 28px; border-radius: 50%;
            line-height: 28px; font-weight: 700; font-size: 13px; text-align: center; }}
  .badge-gold {{ background: #FFD700; color: #333; }}
  .badge-silver {{ background: #C0C0C0; color: #333; }}
  .badge-bronze {{ background: #CD7F32; color: white; }}
  .status {{ display: inline-block; padding: 3px 10px; border-radius: 12px;
             font-size: 12px; font-weight: 600; }}
  .status.good {{ background: #E8F5E9; color: #2E7D32; }}
  .status.warn {{ background: #FFF9C4; color: #F57F17; }}
  .status.danger {{ background: #FFCDD2; color: #C62828; }}
  .footer {{ text-align: center; color: #aaa; font-size: 12px; padding: 16px; }}
</style>
</head>
<body>
<div class="header">
  <h1>Amazon セラーアナリティクス レポート</h1>
  <p>生成日時: {self.generated_at} | セラーID: A2B1UA7OXUW3IW | マーケット: 日本 (JP)</p>
</div>
<div class="container">
  <div class="section">
    <h2>KPI サマリー</h2>
    {kpi_cards}
  </div>
  <div class="section">
    <h2>月次売上推移</h2>
    <table>
      <thead><tr>
        <th>年月</th><th>売上</th><th>注文数</th><th>平均注文額</th><th>前月比成長率</th>
      </tr></thead>
      <tbody>{monthly_rows}</tbody>
    </table>
  </div>
  <div class="section">
    <h2>商品別売上ランキング</h2>
    <table>
      <thead><tr>
        <th>順位</th><th>商品名</th><th>ASIN</th><th>売上</th><th>販売数</th><th>売上シェア</th>
      </tr></thead>
      <tbody>{product_rows}</tbody>
    </table>
  </div>
  <div class="section">
    <h2>在庫状況</h2>
    <table>
      <thead><tr>
        <th>商品名</th><th>SKU</th><th>販売可能在庫</th><th>入荷予定</th>
        <th>在庫状態</th><th>在庫日数</th><th>補充推奨数</th>
      </tr></thead>
      <tbody>{inventory_rows}</tbody>
    </table>
  </div>
</div>
<div class="footer">Amazon Seller Analytics | Powered by SP-API</div>
</body>
</html>"""

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)
        return filepath


# シングルトンインスタンス
report_gen = ReportGenerator()
