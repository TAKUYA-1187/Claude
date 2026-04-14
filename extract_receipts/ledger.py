"""
独自仕訳帳 Excel 生成モジュール（MoneyForward が使えない場合の fallback）

青色申告 複式簿記に対応した仕訳帳と月別集計シートを生成する。
"""
import logging
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .parsers.base import Receipt

logger = logging.getLogger(__name__)

# 配色
COLOR_HEADER = "2C5282"      # 濃い青
COLOR_SUBHEADER = "4A90D9"   # 中間の青
COLOR_ALT_ROW = "EBF4FF"     # 薄い青（交互行）
COLOR_TOTAL = "FFF9C4"       # 薄い黄（合計行）

THIN = Side(style="thin", color="CCCCCC")
MEDIUM = Side(style="medium", color="2C5282")


def _thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _medium_border():
    return Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def _header_fill(color: str = COLOR_HEADER):
    return PatternFill(fill_type="solid", fgColor=color)


def _alt_fill():
    return PatternFill(fill_type="solid", fgColor=COLOR_ALT_ROW)


def _total_fill():
    return PatternFill(fill_type="solid", fgColor=COLOR_TOTAL)


def _build_journal_rows(receipt: Receipt) -> List[Tuple]:
    """Receipt 1件を仕訳行タプルのリストに変換する"""
    rows = []
    store_label = {
        "yahoo": "Yahoo ショッピング",
        "rakuten": "楽天市場",
        "apple": "Apple Store",
    }.get(receipt.store_type, receipt.store_name)

    description = f"{store_label} 注文:{receipt.order_id}"
    if receipt.items:
        description += f" {receipt.items[0].name[:20]}"

    purchase_amount = receipt.total - receipt.shipping_cost

    if receipt.store_type in ("yahoo", "rakuten"):
        debit = "仕入高"
    else:
        debit = "消耗品費" if purchase_amount < Decimal("100000") else "備品"

    if purchase_amount > 0:
        rows.append((
            receipt.order_date,
            debit,
            "未払金",
            "課税仕入10%",
            int(purchase_amount),
            description,
            receipt.order_id,
        ))

    if receipt.shipping_cost > 0:
        rows.append((
            receipt.order_date,
            "荷造運賃",
            "未払金",
            "課税仕入10%",
            int(receipt.shipping_cost),
            f"送料 注文:{receipt.order_id}",
            receipt.order_id,
        ))

    return rows


def _write_journal_sheet(wb: Workbook, receipts: List[Receipt]) -> None:
    """仕訳帳シートを作成する"""
    ws = wb.active
    ws.title = "仕訳帳"

    headers = ["取引日", "借方科目", "貸方科目", "税区分", "金額（税込）", "摘要", "注文番号"]
    col_widths = [14, 14, 12, 16, 16, 50, 22]

    # ヘッダー行
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="游ゴシック", bold=True, color="FFFFFF", size=10)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    # データ行
    all_rows = []
    for receipt in receipts:
        try:
            all_rows.extend(_build_journal_rows(receipt))
        except Exception as e:
            logger.warning(f"仕訳スキップ (UID {receipt.email_uid}): {e}")

    # 日付でソート
    all_rows.sort(key=lambda r: r[0])

    total_amount = 0
    for row_idx, row_data in enumerate(all_rows, 2):
        is_alt = (row_idx % 2 == 0)
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(value, date):
                cell.value = value
                cell.number_format = "YYYY/MM/DD"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(value, int):
                cell.value = value
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                total_amount += value
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="游ゴシック", size=10)
            cell.border = _thin_border()
            if is_alt:
                cell.fill = _alt_fill()

    # 合計行
    total_row = len(all_rows) + 2
    ws.cell(row=total_row, column=1, value="合　計").font = Font(name="游ゴシック", bold=True, size=10)
    total_cell = ws.cell(row=total_row, column=5, value=total_amount)
    total_cell.number_format = "#,##0"
    total_cell.font = Font(name="游ゴシック", bold=True, size=11)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _total_fill()
        cell.border = _thin_border()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_rows) + 1}"


def _write_summary_sheet(wb: Workbook, receipts: List[Receipt]) -> None:
    """月別集計シートを作成する"""
    ws = wb.create_sheet(title="月別集計")

    # 月別集計を計算
    monthly: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for receipt in receipts:
        month_key = receipt.order_date.strftime("%Y年%m月")
        purchase = int(receipt.total - receipt.shipping_cost)
        shipping = int(receipt.shipping_cost)
        monthly[month_key]["仕入高・消耗品費"] += purchase
        monthly[month_key]["荷造運賃"] += shipping
        monthly[month_key]["合計"] += purchase + shipping

    headers = ["月", "仕入高・消耗品費", "荷造運賃", "合計"]
    col_widths = [14, 18, 14, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="游ゴシック", bold=True, color="FFFFFF", size=10)
        cell.fill = _header_fill(COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    sorted_months = sorted(monthly.keys())
    grand_totals = defaultdict(int)

    for row_idx, month in enumerate(sorted_months, 2):
        is_alt = (row_idx % 2 == 0)
        data = monthly[month]
        row_vals = [
            month,
            data["仕入高・消耗品費"],
            data["荷造運賃"],
            data["合計"],
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="游ゴシック", size=10)
            cell.border = _thin_border()
            if isinstance(val, int):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col > 1:
                    grand_totals[headers[col - 1]] += val
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_alt:
                cell.fill = _alt_fill()

    # 年間合計行
    total_row = len(sorted_months) + 2
    ws.cell(row=total_row, column=1, value="年間合計").font = Font(
        name="游ゴシック", bold=True, size=10
    )
    for col, h in enumerate(headers[1:], 2):
        cell = ws.cell(row=total_row, column=col, value=grand_totals[h])
        cell.number_format = "#,##0"
        cell.font = Font(name="游ゴシック", bold=True, size=11)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = _total_fill()
        cell.border = _thin_border()
    ws.cell(row=total_row, column=1).fill = _total_fill()
    ws.cell(row=total_row, column=1).border = _thin_border()


def export_to_excel(
    receipts: List[Receipt],
    output_path: Optional[Path] = None,
    dry_run: bool = False,
) -> Path:
    """
    Receipt リストを Excel 仕訳帳に書き出す。

    Args:
        receipts: 領収書データのリスト
        output_path: 出力先パス（省略時は自動生成）
        dry_run: True の場合はパスを返すのみでファイルを作成しない

    Returns:
        出力した Excel ファイルのパス
    """
    if output_path is None:
        today = date.today().strftime("%Y%m%d")
        output_path = Path(f"ledger_仕訳帳_{today}.xlsx")

    if dry_run:
        logger.info(f"[DRY RUN] Excel 仕訳帳出力: {output_path}")
        return output_path

    wb = Workbook()
    _write_journal_sheet(wb, receipts)
    _write_summary_sheet(wb, receipts)
    wb.save(output_path)

    logger.info(f"Excel 仕訳帳出力完了: {output_path} ({len(receipts)} 件)")
    return output_path
