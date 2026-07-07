"""Amazon広告の検索用語レポート / ターゲティングレポートの読み込み。

管理画面（広告レポート）からダウンロードした .xlsx / .csv をそのまま渡せる。
日本語・英語どちらの列名でも読めるよう、列名はエイリアス表で正規化する。
"""

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class Row:
    campaign: str = ""
    ad_group: str = ""
    targeting: str = ""          # ターゲティング（キーワード or ASIN or auto区分）
    match_type: str = ""
    search_term: str = ""        # 検索用語レポートのみ。ターゲティングレポートでは空
    impressions: int = 0
    clicks: int = 0
    spend: float = 0.0
    sales: float = 0.0
    orders: int = 0
    raw: dict = field(default_factory=dict)

    @property
    def cpc(self) -> float:
        return self.spend / self.clicks if self.clicks else 0.0

    @property
    def acos(self) -> float | None:
        """売上ゼロのときは None（÷0を避ける）。"""
        return self.spend / self.sales if self.sales else None

    @property
    def value_per_click(self) -> float:
        return self.sales / self.clicks if self.clicks else 0.0


# 正規化後の列名に含まれていれば一致とみなす部分文字列。
# 上から順に評価し、最初に一致したフィールドへ割り当てる。
_FIELD_PATTERNS: list[tuple[str, list[str]]] = [
    ("campaign", ["キャンペーン名", "campaignname"]),
    ("ad_group", ["広告グループ名", "adgroupname"]),
    ("match_type", ["マッチタイプ", "matchtype"]),
    ("search_term", ["検索キーワード", "検索用語", "searchterm"]),
    ("targeting", ["ターゲティング", "targeting"]),
    ("impressions", ["インプレッション", "impression"]),
    ("clicks", ["クリック数", "clicks"]),
    ("spend", ["費用", "spend", "cost"]),
    ("orders", ["注文数", "orders"]),
    ("sales", ["売上", "sales"]),
]

_INT_FIELDS = {"impressions", "clicks", "orders"}
_FLOAT_FIELDS = {"spend", "sales"}


def _normalize(name: str) -> str:
    """列名から空白・記号を除いて小文字化する。"""
    return re.sub(r"[\s()（）#￥$%・,、。「」【】-]", "", str(name)).lower()


def _map_columns(header: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    used: set[str] = set()
    for idx, col in enumerate(header):
        norm = _normalize(col)
        if not norm:
            continue
        for fld, patterns in _FIELD_PATTERNS:
            if fld in used:
                continue
            if any(p in norm for p in patterns):
                mapping[idx] = fld
                used.add(fld)
                break
    return mapping


def _to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[￥$,\s%]", "", str(value))
    if not s or s in {"-", "—"}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _rows_from_table(header: list[str], data_rows) -> list[Row]:
    mapping = _map_columns(header)
    if "campaign" not in mapping.values() or "clicks" not in mapping.values():
        raise ValueError(
            "列名を認識できませんでした。Amazon広告の検索用語レポートまたは"
            "ターゲティングレポートを、加工せずそのまま渡してください。"
            f"（検出したヘッダー: {header}）"
        )
    rows = []
    for values in data_rows:
        row = Row()
        for idx, fld in mapping.items():
            if idx >= len(values):
                continue
            value = values[idx]
            if fld in _INT_FIELDS:
                setattr(row, fld, int(_to_number(value)))
            elif fld in _FLOAT_FIELDS:
                setattr(row, fld, _to_number(value))
            else:
                setattr(row, fld, "" if value is None else str(value).strip())
        row.raw = dict(zip(header, values))
        if row.campaign:
            rows.append(row)
    return rows


def load_report(path: str | Path) -> list[Row]:
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        return _rows_from_table(header, it)

    # CSV: Amazon JPのエクスポートは UTF-8(BOM付き) か Shift_JIS のどちらか
    for encoding in ("utf-8-sig", "cp932"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                reader = csv.reader(f)
                header = next(reader)
                return _rows_from_table(header, list(reader))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"{path} の文字コードを判定できませんでした。")
